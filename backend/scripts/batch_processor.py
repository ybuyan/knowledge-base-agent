#!/usr/bin/env python
"""
Batch Processor
Process documents in batch mode with parallel processing support.
"""

import os
import sys
import asyncio
import logging
import hashlib
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.core.file_tracker import FileTracker
from app.core.chroma import get_documents_collection, delete_document_vectors
from app.core.embeddings import get_embeddings
from app.models.document import DocumentDB, DocumentStatusModel
from app.core.constants import DocumentConstants

logger = logging.getLogger(__name__)


class BatchProcessor:
    """Process documents in batch"""
    
    # 使用统一的扩展名映射
    SUPPORTED_EXTENSIONS = DocumentConstants.EXTENSION_TO_TYPE
    
    def __init__(self, tracker: FileTracker = None, db_path: str = None):
        if tracker:
            self.tracker = tracker
        elif db_path:
            self.tracker = FileTracker(db_path)
        else:
            self.tracker = FileTracker("data/watch_state.db")
        
        self.embeddings = get_embeddings()
        self.collection = get_documents_collection()
    
    def _get_file_type(self, file_path: str) -> Optional[str]:
        """Get file type from extension"""
        ext = os.path.splitext(file_path)[1].lower()
        return self.SUPPORTED_EXTENSIONS.get(ext)
    
    def _parse_document(self, file_path: str) -> str:
        """Parse document and extract text"""
        file_type = self._get_file_type(file_path)
        
        if file_type == 'pdf':
            return self._parse_pdf(file_path)
        elif file_type == 'docx':
            return self._parse_docx(file_path)
        elif file_type == 'txt':
            return self._parse_txt(file_path)
        elif file_type == 'xlsx':
            return self._parse_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_path}")
    
    def _parse_pdf(self, file_path: str) -> str:
        """Parse PDF file"""
        import PyPDF2
        
        filename = os.path.basename(file_path)
        text_parts = []
        
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for i, page in enumerate(reader.pages):
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(f"[{filename} - Page {i+1}]\n{page_text}")
        
        return "\n\n".join(text_parts)
    
    def _parse_docx(self, file_path: str) -> str:
        """Parse DOCX file"""
        from docx import Document
        
        doc = Document(file_path)
        return "\n\n".join([para.text for para in doc.paragraphs if para.text])
    
    def _parse_txt(self, file_path: str) -> str:
        """Parse TXT file"""
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    def _parse_xlsx(self, file_path: str) -> str:
        """Parse Excel file"""
        from openpyxl import load_workbook
        
        filename = os.path.basename(file_path)
        text_parts = []
        
        wb = load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"[{filename} - 工作表: {sheet_name}]")
            
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            
            # 第一行作为表头
            headers = [str(cell) if cell is not None else f"列{i+1}" for i, cell in enumerate(rows[0])]
            text_parts.append(f"表头: {' | '.join(headers)}")
            text_parts.append("")
            
            # 处理数据行 - 每行作为一个完整的记录块
            for row_idx, row in enumerate(rows[1:], 2):
                if not any(cell is not None for cell in row):
                    continue
                
                # 构建键值对格式的行记录
                row_data = []
                for i, (header, cell) in enumerate(zip(headers, row)):
                    if cell is not None:
                        row_data.append(f"{header}: {cell}")
                
                if row_data:
                    # 将整行数据作为一个完整的文本块
                    record_text = f"记录 {row_idx-1}: " + " | ".join(row_data)
                    text_parts.append(record_text)
            
            text_parts.append("")
        
        return "\n".join(text_parts)
    
    def _split_text(self, text: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """Split text into chunks"""
        paragraphs = text.split("\n\n")
        chunks = []
        current_chunk = ""
        
        for para in paragraphs:
            if len(current_chunk) + len(para) <= chunk_size:
                current_chunk += para + "\n\n"
            else:
                if current_chunk:
                    chunks.append(current_chunk.strip())
                current_chunk = para + "\n\n"
        
        if current_chunk:
            chunks.append(current_chunk.strip())
        
        return chunks
    
    def _generate_document_id(self, file_path: str) -> str:
        """Generate unique document ID"""
        return hashlib.md5(file_path.encode()).hexdigest()
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """Process a single file"""
        if not os.path.exists(file_path):
            error_msg = f"File not found: {file_path}"
            logger.error(error_msg)
            self.tracker.mark_failed(file_path, error_msg)
            return {"success": False, "error": error_msg}
        
        # Check if file type is supported
        if not self._get_file_type(file_path):
            error_msg = f"Unsupported file type: {file_path}"
            logger.error(error_msg)
            self.tracker.mark_failed(file_path, error_msg)
            return {"success": False, "error": error_msg}
        
        try:
            # Mark as processing
            self.tracker.mark_processing(file_path)
            
            # Parse document
            logger.info(f"Parsing document: {file_path}")
            text = self._parse_document(file_path)
            
            if not text.strip():
                error_msg = f"Empty document: {file_path}"
                logger.warning(error_msg)
                self.tracker.mark_failed(file_path, error_msg)
                return {"success": False, "error": error_msg}
            
            # Split into chunks
            chunks = self._split_text(text)
            logger.info(f"Split into {len(chunks)} chunks")
            
            # Generate embeddings in batches to avoid 400 error
            logger.info(f"Generating embeddings for {len(chunks)} chunks")
            batch_size = 10  # Process 10 chunks at a time
            embeddings_list = []
            
            # Use synchronous embedding method instead of async
            for i in range(0, len(chunks), batch_size):
                batch = chunks[i:i+batch_size]
                logger.info(f"Processing batch {i//batch_size + 1}/{(len(chunks)-1)//batch_size + 1}")
                batch_embeddings = self.embeddings.embed_documents(batch)
                embeddings_list.extend(batch_embeddings)
            
            # Generate document ID
            document_id = self._generate_document_id(file_path)
            filename = os.path.basename(file_path)
            
            # Store in ChromaDB
            ids = [f"{document_id}_chunk_{i}" for i in range(len(chunks))]
            metadatas = [
                {
                    "document_id": document_id,
                    "document_name": filename,
                    "chunk_index": i,
                    "source_file": file_path
                }
                for i in range(len(chunks))
            ]
            
            self.collection.add(
                ids=ids,
                embeddings=embeddings_list,
                documents=chunks,
                metadatas=metadatas
            )
            
            # Save to MongoDB for API access
            try:
                stat = os.stat(file_path)
                upload_time = datetime.fromtimestamp(stat.st_mtime).isoformat()
                
                doc = DocumentStatusModel(
                    id=document_id,
                    filename=filename,
                    status="READY",
                    size=stat.st_size,
                    uploadTime=upload_time,
                    file_path=file_path,
                    chunk_count=len(chunks)
                )
                DocumentDB.save_sync(doc)
                logger.info(f"Saved document to MongoDB: {filename}")
            except Exception as db_error:
                logger.warning(f"Failed to save to MongoDB: {db_error}")
            
            # Mark as done
            self.tracker.mark_done(file_path, len(chunks), document_id)
            
            logger.info(f"Successfully processed: {file_path} ({len(chunks)} chunks)")
            
            return {
                "success": True,
                "document_id": document_id,
                "chunk_count": len(chunks),
                "file_path": file_path
            }
            
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Error processing {file_path}: {error_msg}")
            self.tracker.mark_failed(file_path, error_msg)
            return {"success": False, "error": error_msg}
    
    def process_batch(self, file_paths: List[str]) -> Dict[str, Any]:
        """Process multiple files"""
        results = {
            "total": len(file_paths),
            "success": 0,
            "failed": 0,
            "details": []
        }
        
        for file_path in file_paths:
            result = self.process_file(file_path)
            results["details"].append(result)
            
            if result["success"]:
                results["success"] += 1
            else:
                results["failed"] += 1
        
        return results
    
    def process_pending(self, limit: int = 100) -> Dict[str, Any]:
        """Process all pending files"""
        pending_files = self.tracker.get_pending_files(limit)
        
        if not pending_files:
            logger.info("No pending files to process")
            return {"total": 0, "success": 0, "failed": 0, "details": []}
        
        logger.info(f"Processing {len(pending_files)} pending files")
        
        file_paths = [f["file_path"] for f in pending_files]
        return self.process_batch(file_paths)
    
    def retry_failed(self, limit: int = 100) -> Dict[str, Any]:
        """Retry failed files"""
        failed_files = self.tracker.get_failed_files(limit)
        
        if not failed_files:
            logger.info("No failed files to retry")
            return {"total": 0, "success": 0, "failed": 0, "details": []}
        
        logger.info(f"Retrying {len(failed_files)} failed files")
        
        # Mark all as pending
        for f in failed_files:
            self.tracker.mark_for_reprocess(f["file_path"])
        
        return self.process_pending(limit)
    
    def get_stats(self) -> Dict[str, Any]:
        """Get processing statistics"""
        tracker_stats = self.tracker.get_stats()
        chroma_count = self.collection.count()
        
        return {
            **tracker_stats,
            "vectors_in_db": chroma_count
        }


async def main():
    """Main entry point for batch processing"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Batch Document Processor")
    parser.add_argument("--dir", "-d", help="Directory to process")
    parser.add_argument("--file", "-f", help="Single file to process")
    parser.add_argument("--pending", "-p", action="store_true", help="Process pending files")
    parser.add_argument("--retry", "-r", action="store_true", help="Retry failed files")
    parser.add_argument("--stats", "-s", action="store_true", help="Show statistics")
    parser.add_argument("--limit", "-l", type=int, default=100, help="Limit number of files")
    parser.add_argument("--extensions", "-e", nargs="+", default=DocumentConstants.SUPPORTED_EXTENSIONS)
    
    args = parser.parse_args()
    
    processor = BatchProcessor()
    
    if args.stats:
        stats = processor.get_stats()
        print("\n=== Processing Statistics ===")
        for key, value in stats.items():
            print(f"  {key}: {value}")
        return
    
    if args.file:
        result = processor.process_file(args.file)
        print(f"\nResult: {result}")
        return
    
    if args.pending:
        results = processor.process_pending(args.limit)
        print(f"\nProcessed: {results['success']}/{results['total']} successful")
        return
    
    if args.retry:
        results = processor.retry_failed(args.limit)
        print(f"\nRetried: {results['success']}/{results['total']} successful")
        return
    
    if args.dir:
        # Scan directory and process
        extensions = set(ext.lower() for ext in args.extensions)
        file_paths = []
        
        for root, dirs, files in os.walk(args.dir):
            for file in files:
                ext = os.path.splitext(file)[1].lower()
                if ext in extensions:
                    file_paths.append(os.path.join(root, file))
        
        print(f"\nFound {len(file_paths)} files to process")
        
        # Register files
        for fp in file_paths:
            if not processor.tracker.file_exists(fp):
                processor.tracker.register_file(fp)
        
        # Process
        results = processor.process_batch(file_paths)
        print(f"\nProcessed: {results['success']}/{results['total']} successful")
        return
    
    parser.print_help()


if __name__ == "__main__":
    asyncio.run(main())
