#!/usr/bin/env python
"""
Test XLSX parsing in batch processor
"""

import os
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> str:
    """Parse Excel file - same logic as batch_processor"""
    filename = os.path.basename(file_path)
    text_parts = []
    
    wb = load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_parts.append(f"[{filename} - 工作表: {sheet_name}]")
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        
        # 第一行作为表头
        headers = [str(cell) if cell is not None else f"列{i+1}" for i, cell in enumerate(rows[0])]
        text_parts.append(f"表头: {' | '.join(headers)}")
        text_parts.append("")
        
        # 处理数据行 - 每行作为一个完整的记录块
        for row_idx, row in enumerate(rows[1:], 2):
            if not any(cell is not None for cell in row):
                continue
            
            # 构建键值对格式的行记录
            row_data = []
            for i, (header, cell) in enumerate(zip(headers, row)):
                if cell is not None:
                    row_data.append(f"{header}: {cell}")
            
            if row_data:
                # 将整行数据作为一个完整的文本块
                record_text = f"记录 {row_idx-1}: " + " | ".join(row_data)
                text_parts.append(record_text)
        
        text_parts.append("")
    
    return "\n".join(text_parts)


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Test XLSX parsing")
    parser.add_argument("file", help="XLSX file to parse")
    
    args = parser.parse_args()
    
    if not os.path.exists(args.file):
        print(f"Error: File not found: {args.file}")
        sys.exit(1)
    
    try:
        result = parse_xlsx(args.file)
        print("=== Parsed Content ===")
        print(result)
        print("\n=== Success ===")
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
